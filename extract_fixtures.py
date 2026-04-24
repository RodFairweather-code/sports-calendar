import openpyxl, json
from datetime import datetime

def fmt(dt):
    if isinstance(dt, datetime):
        return dt.strftime('%Y-%m-%dT%H:%M')
    return None

# ── EPL ──────────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(r'D:\Documents\000 Merapar\Sky Sports Phase 2\Sports schedules import\EPL_2025_2026_Full_Season_Airtable.xlsx')
ws = wb['EPL Schedule']
rows = list(ws.iter_rows(values_only=True))

epl = []
for r in rows[1:]:
    rnd, sport, league, dt_str, stadium, home, away = r[:7]
    if not home or not away:
        continue
    if isinstance(dt_str, str):
        try:
            start = datetime.strptime(dt_str, '%d/%m/%Y %H:%M').strftime('%Y-%m-%dT%H:%M')
        except:
            start = None
    elif isinstance(dt_str, datetime):
        start = dt_str.strftime('%Y-%m-%dT%H:%M')
    else:
        start = None
    epl.append({'round': rnd, 'start': start, 'stadium': stadium,
                'homeTeam': home, 'awayTeam': away,
                'homeScore': None, 'awayScore': None})

with open(r'C:\Users\rodfa\claude-cli\sports-calendar\src\data\fixtures\epl_2025.json', 'w') as f:
    json.dump(epl, f, indent=2)
print(f'EPL: {len(epl)} fixtures written')

# ── Championship ─────────────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(r'D:\Documents\000 Merapar\Sky Sports Phase 2\Sports schedules import\Raw download Championship_League_Interactive_Table_2025-26_rev1.5.xlsx')
ws2 = wb2['fixtures_tmp']
rows2 = list(ws2.iter_rows(values_only=True))
# header: Match Day, MatchDate, Match Time(NY), Match Time(London), ..., Stadium, Home Team, Home Goal, Away Goal, Away Team

champ = []
for r in rows2[1:]:
    rnd = r[0]
    dt_london = r[3]
    stadium = r[8]
    home = r[9]
    home_score = r[10]
    away_score = r[11]
    away = r[12]
    if not home or not away:
        continue
    start = fmt(dt_london)
    champ.append({'round': rnd, 'start': start, 'stadium': stadium,
                  'homeTeam': home, 'awayTeam': away,
                  'homeScore': home_score, 'awayScore': away_score})

with open(r'C:\Users\rodfa\claude-cli\sports-calendar\src\data\fixtures\championship_2025.json', 'w') as f:
    json.dump(champ, f, indent=2)
print(f'Championship: {len(champ)} fixtures written')

# ── F1 ───────────────────────────────────────────────────────────────────────
wb3 = openpyxl.load_workbook(r'D:\Documents\000 Merapar\Sky Sports Phase 2\Sports schedules import\F1_2026_Full_Schedule_Separated.xlsx')
ws3 = wb3['Sheet1']
rows3 = list(ws3.iter_rows(values_only=True))
# Race #, Grand Prix, Location, Circuit, Sprint Weekend, Session, Local Date, Local Time, UK Date, UK Time

f1 = []
for r in rows3[1:]:
    race_num, gp, location, circuit, sprint, session, local_date, local_time, uk_date, uk_time = r[:10]
    if not gp or not session:
        continue
    if uk_date and uk_time:
        start = f'{uk_date}T{uk_time}'
    elif uk_date:
        start = str(uk_date)
    else:
        start = None
    f1.append({'raceNum': race_num, 'grandPrix': gp, 'location': location,
               'circuit': circuit, 'sprintWeekend': sprint == 'Y',
               'session': session, 'start': start})

with open(r'C:\Users\rodfa\claude-cli\sports-calendar\src\data\fixtures\f1_2026.json', 'w') as f:
    json.dump(f1, f, indent=2)
print(f'F1: {len(f1)} sessions written')
